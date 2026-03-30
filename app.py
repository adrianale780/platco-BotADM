
import streamlit as st
import io
import requests
import openpyxl
from datetime import datetime, timedelta
import os
import re
import unicodedata



# ==========================================
# 🧠 UTILIDADES GENERALES
# ==========================================
def normalizar_texto(texto):
    if not texto: return ""
    texto = str(texto).upper().strip()
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def obtener_hoja_flexible(wb, nombre_buscado):
    objetivo = normalizar_texto(nombre_buscado)
    for sheet_name in wb.sheetnames:
        if normalizar_texto(sheet_name) == objetivo:
            return wb[sheet_name]
    return None

def cargar_diccionario_cuentas():
    reglas = []
    ruta_diccionario = os.path.join(os.path.dirname(__file__), "diccionario.xlsx")
    if not os.path.exists(ruta_diccionario): return []
    try:
        wb = openpyxl.load_workbook(ruta_diccionario, read_only=True, data_only=True)
        ws = obtener_hoja_flexible(wb, "CATEGORIA") or wb.active
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0] and fila[1]: 
                reglas.append((str(fila[0]).strip().upper(), str(fila[1]).strip()))
        wb.close()
        reglas.sort(key=lambda x: len(x[0]), reverse=True)
        return reglas
    except: return []

def cargar_diccionario_areas():
    reglas = []
    ruta_diccionario = os.path.join(os.path.dirname(__file__), "diccionario.xlsx")
    if not os.path.exists(ruta_diccionario): return []
    try:
        wb = openpyxl.load_workbook(ruta_diccionario, read_only=True, data_only=True)
        ws = obtener_hoja_flexible(wb, "AREA")
        if ws:
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if fila[0] and fila[1]: 
                    reglas.append((str(fila[0]).strip().upper(), str(fila[1]).strip()))
        wb.close()
        reglas.sort(key=lambda x: len(x[0]), reverse=True)
        return reglas
    except: return []

# ==========================================
# 🧠 CÁLCULO DE TASAS HISTÓRICAS (CON CLAVE PREMIUM)
# ==========================================
def cargar_tasas_historicas(callback_log):
    memoria_tasas = {}
    URL_HISTORICO = "https://api.dolarvzla.com/public/exchange-rate/list"
    
    # TU CLAVE DE ACCESO
    MI_CLAVE = "b985611cba55720f44cad13d1cb8e16e8963e3deb25950ecbcd646024b6f8934"
    
    # Preparamos el pase de entrada (Header)
    headers = {
        'x-dolarvzla-key': MI_CLAVE
    }
    
    callback_log(f"🌍 Conectando al histórico oficial (Con llave)...")

    try:
        # Enviamos la petición CON la clave
        response = requests.get(URL_HISTORICO, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            lista_datos = []

            if isinstance(data, dict) and 'rates' in data:
                lista_datos = data['rates']
            elif isinstance(data, list):
                lista_datos = data
            
            count = 0
            for item in lista_datos:
                fecha = item.get('date')
                precio = item.get('usd')

                if fecha and precio:
                    fecha_limpia = str(fecha)[:10]
                    memoria_tasas[fecha_limpia] = float(precio)
                    count += 1
            
            callback_log(f"   ✅ Acceso concedido: {count} fechas históricas cargadas.")
        else:
            # Si la clave falla o expira, avisamos pero no detenemos el programa
            callback_log(f"⚠️ Error de acceso al histórico ({response.status_code}). Usando solo tasa de hoy.")

    except Exception as e:
        callback_log(f"⚠️ Sin conexión al histórico: {str(e)}")

    return memoria_tasas

def formatear_fecha_para_api(valor_celda):
    if not valor_celda: return None
    try:
        if isinstance(valor_celda, datetime): return valor_celda.strftime("%Y-%m-%d")
        texto = str(valor_celda).strip()
        formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y"]
        for fmt in formatos:
            try: return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
            except: continue
    except: return None
    return None

def buscar_tasa_inteligente(fecha_str, memoria_tasas):
    if not fecha_str: return 0
    if fecha_str in memoria_tasas: return memoria_tasas[fecha_str]
    try:
        dt_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
        for dias_atras in range(1, 6):
            fecha_b = dt_obj - timedelta(days=dias_atras)
            f_key = fecha_b.strftime("%Y-%m-%d")
            if f_key in memoria_tasas: return memoria_tasas[f_key]
    except: pass
    return 0

def limpiar_numero(valor):
    if valor is None: return 0
    if isinstance(valor, (int, float)): return float(valor)
    texto = str(valor).strip()
    if not texto or texto == "-": return 0
    try:
        return float(texto.replace(".", "").replace(",", "."))
    except: return 0

def obtener_nombre_mes_es(mes_num):
    meses = {1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
             7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"}
    return meses.get(mes_num, "")

from openpyxl.utils import get_column_letter
from datetime import datetime
import re


# ==========================================
# 🛡️ CÓDIGO ORIGINAL (ANOCHE): RESUMEN SEMANAL
# ==========================================
# Solo procesa Ingresos. Solo usa DATA BS. No toca Excedentes.

def procesar_resumen_semanal(wb, ruta_excel, callback_log):
    import re
    import openpyxl
    from openpyxl.utils import get_column_letter
    import unicodedata

    # 1. Utilidades Internas
    def normalizar_texto_local(texto):
        if not texto: return ""
        texto = str(texto).upper().strip()
        return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

    def limpiar_numero(valor):
        if valor is None: return 0
        if isinstance(valor, (int, float)): return float(valor)
        txt = str(valor).strip().upper().replace("BS", "").replace("USD", "").replace(" ", "")
        es_neg = False
        if "(" in txt and ")" in txt:
            es_neg = True
            txt = txt.replace("(", "").replace(")", "")
        if "." in txt and "," in txt: txt = txt.replace(".", "").replace(",", ".")
        elif "," in txt: txt = txt.replace(",", ".")
        try: return -float(txt) if es_neg else float(txt)
        except: return 0

    ws_resumen = None
    ws_data = None
    for sheet in wb.sheetnames:
        norm_sheet = normalizar_texto_local(sheet)
        if "FLUJO DE CAJA" in norm_sheet:
            ws_resumen = wb[sheet]
        if "DATA BS" in norm_sheet:
            ws_data = wb[sheet]
            
    if not ws_resumen or not ws_data:
        callback_log("⚠️ Error: Faltan hojas (Resumen o Data BS).")
        return 0

    callback_log("📅 Procesando Resumen Semanal en USD...")

    # --- LECTURA SEGURA: VALORES REALES (IGNORANDO FÓRMULAS) ---
    callback_log("⏳ Extrayendo datos de semanas y dólares (modo lectura)...")
    data_bs_lista = []
    try:
        # Abrimos el archivo en modo "solo resultados"
        wb_lectura = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
        ws_data_L = None
        for sheet in wb_lectura.sheetnames:
            if "DATA BS" in normalizar_texto_local(sheet):
                ws_data_L = wb_lectura[sheet]
                break
        
        if ws_data_L:
            # Pasamos todo a la memoria RAM rapidísimo
            for row in ws_data_L.iter_rows(min_row=4, values_only=True):
                data_bs_lista.append(row)
        wb_lectura.close()
    except Exception as e:
        callback_log(f"❌ Error leyendo DATA BS: {str(e)}")
        return 0

    # 3. DETECTAR SEMANAS EN EL ENCABEZADO DE RESUMEN
    fila_encabezado = 0
    for r in range(1, 15):
        for c in range(1, 20): 
            val = str(ws_resumen.cell(row=r, column=c).value).upper()
            if "SEMANA" in val and any(char.isdigit() for char in val):
                fila_encabezado = r
                break
        if fila_encabezado > 0: break
    
    if fila_encabezado == 0: 
        callback_log("⚠️ No encontré encabezados como 'SEMANA 1' en la hoja Resumen.")
        return 0

    semanas_config = [] 
    for col in range(2, ws_resumen.max_column + 1):
        texto = str(ws_resumen.cell(row=fila_encabezado, column=col).value).upper()
        match = re.search(r"SEMANA\s*(\d+)", texto)
        if match:
            num_semana = int(match.group(1))
            sub = normalizar_texto_local(ws_resumen.cell(row=fila_encabezado+1, column=col).value)
            col_est, col_act = 0, 0
            
            if "ACTUALIZACION" in sub or "DIARIA" in sub:
                col_act, col_est = col, col - 1
            elif "ESTIMADO" in sub or "INICIAL" in sub:
                col_est, col_act = col, col + 1
            
            if col_est > 0 and col_act > 0:
                semanas_config.append({
                    'semana': num_semana,
                    'col_est': col_est, 'col_act': col_act
                })

    if not semanas_config:
        callback_log("⚠️ No detecté columnas válidas de Estimado/Actualización.")
        return 0

    # 4. SUMAR REAL (DATA BS - LEYENDO DESDE LA MEMORIA)
    acumulados_real = {cfg['semana']: {} for cfg in semanas_config}
    
    for row in data_bs_lista:
        if len(row) < 17: continue # Aseguramos que la fila llega hasta la columna Q
        
        # Columna Q (Índice 16 en memoria)
        val_semana = row[16]
        if not val_semana: continue
        
        texto_semana = str(val_semana).upper()
        match_sem = re.search(r"SEMANA\s*(\d+)", texto_semana)
        if not match_sem: continue
        semana_del_dato = int(match_sem.group(1))

        # Columna M (Índice 12 en memoria)
        cta = normalizar_texto_local(row[12])
        if not cta: continue
        
        # Columna G (Índice 6 en memoria) - BS
        monto = limpiar_numero(row[6])

        if monto != 0 and semana_del_dato in acumulados_real:
            if cta not in acumulados_real[semana_del_dato]: 
                acumulados_real[semana_del_dato][cta] = 0
            acumulados_real[semana_del_dato][cta] += monto

    # 5. ESCRIBIR EN RESUMEN
    cuentas_ingresos = ["CONTINUIDAD OPERATIVA", "SIMCARD", "ALIADOS COMERCIALES", "BANCO MERCANTIL 20% TX", "BANCO PROVINCIAL 20% TX"]
    start_row = fila_encabezado + 2 
    cambios = 0
    
    for r in range(start_row, ws_resumen.max_row + 1):
        # Escaneamos las primeras 3 columnas (A, B, C) por si el texto está indentado o movido
        nombre = ""
        for c in range(1, 4):
            val_cta = ws_resumen.cell(row=r, column=c).value
            if val_cta:
                nombre_temp = normalizar_texto_local(val_cta)
                if any(cta_valida in nombre_temp for cta_valida in cuentas_ingresos):
                    nombre = nombre_temp
                    break
        
        if not nombre: continue # Si no encuentra cuenta válida, salta a la siguiente fila

        for cfg in semanas_config:
            num_sem = cfg['semana']
            monto_real = 0
            
            for k, v in acumulados_real[num_sem].items():
                if nombre in k or k in nombre: 
                    monto_real += v
            
            valor_estimado = ws_resumen.cell(row=r, column=cfg['col_est']).value
            
            # Si hay estimado manual o hubo movimientos reales, reescribimos la fórmula
            if valor_estimado or monto_real != 0:
                letra_est = get_column_letter(cfg['col_est'])
                monto_formateado = round(monto_real, 2)
                
                # FÓRMULA: =ESTIMADO - REAL
                ws_resumen.cell(row=r, column=cfg['col_act']).value = f"={letra_est}{r}-{monto_formateado}"
                ws_resumen.cell(row=r, column=cfg['col_act']).number_format = '#,##0.00'
                cambios += 1

    return cambios


def procesar_conciliacion_compleja(wb, ruta_excel, callback_log):
    from datetime import datetime
    import openpyxl

    # 1. Limpiador de números
    def limpiar_venezuela(valor):
        if not valor: return 0
        if isinstance(valor, (int, float)): return float(valor)
        txt = str(valor).strip().upper().replace("BS", "").replace("USD", "").replace(" ", "")
        es_neg = False
        if "(" in txt and ")" in txt:
            es_neg = True
            txt = txt.replace("(", "").replace(")", "")
        if "." in txt and "," in txt: txt = txt.replace(".", "").replace(",", ".")
        elif "," in txt: txt = txt.replace(",", ".")
        try: return -float(txt) if es_neg else float(txt)
        except: return 0

    # 2. Extractor de Mes Inteligente
    def extraer_mes_inteligente(valor_fecha):
        if not valor_fecha: return ""
        if isinstance(valor_fecha, datetime): return obtener_nombre_mes_es(valor_fecha.month)
        
        txt = str(valor_fecha).strip().upper()
        
        for i in range(1, 13):
            nombre = obtener_nombre_mes_es(i)
            if nombre in txt: return nombre
            
        formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y", "%m-%Y", "%m/%Y"]
        txt_solo_fecha = txt.split(" ")[0]
        for fmt in formatos:
            try: return obtener_nombre_mes_es(datetime.strptime(txt_solo_fecha, fmt).month)
            except: pass
            
        abreviaturas = {"ENE": "ENERO", "FEB": "FEBRERO", "MAR": "MARZO", "ABR": "ABRIL", "MAY": "MAYO", "JUN": "JUNIO", "JUL": "JULIO", "AGO": "AGOSTO", "SEP": "SEPTIEMBRE", "OCT": "OCTUBRE", "NOV": "NOVIEMBRE", "DIC": "DICIEMBRE"}
        for abrv, completo in abreviaturas.items():
            if abrv in txt: return completo

        return normalizar_texto(txt)

    # 3. Matcher de Bancos Flexible
    def bancos_coinciden(b1, b2):
        if not b1 or not b2: return False
        if "PROVINCIAL" in b1 and "PROVINCIAL" in b2: return True
        if "MERCANTIL" in b1 and "MERCANTIL" in b2: return True
        return (b1 in b2) or (b2 in b1)

    ws_apartados = obtener_hoja_flexible(wb, "APARTADOS")
    ws_data = obtener_hoja_flexible(wb, "DATA BS")

    if not ws_apartados or not ws_data:
        callback_log("⚠️ Error: Faltan hojas base.")
        return 0 

    # --- LECTURA SEGURA: FÓRMULAS DIRECTO DEL ARCHIVO ORIGINAL ---
    callback_log("⏳ Extrayendo fórmulas de Excedentes (sin bloqueos)...")
    excedentes_lista = []
    try:
        # Lee el archivo directamente del disco (rápido y a prueba de errores)
        wb_lectura = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
        ws_exc_L = obtener_hoja_flexible(wb_lectura, "MANEJO EXCEDENTE")
        if ws_exc_L:
            for row in ws_exc_L.iter_rows(min_row=4, max_row=2000, values_only=True):
                excedentes_lista.append(row)
        wb_lectura.close() # Cierra inmediatamente
    except Exception as e:
        callback_log(f"❌ Error leyendo fórmulas: {str(e)}")
        return 0

    cambios = 0
    
    # Recorremos hoja APARTADOS
    for fila in range(4, ws_apartados.max_row + 1):
        celda_banco = ws_apartados.cell(row=fila, column=2)
        celda_monto = ws_apartados.cell(row=fila, column=3)
        celda_concepto = ws_apartados.cell(row=fila, column=4)
        celda_mes = ws_apartados.cell(row=fila, column=5)
        
        concepto = normalizar_texto(celda_concepto.value) if celda_concepto.value else ""
        
        if "ESPECIALIZAD" in concepto:
            banco_objetivo = normalizar_texto(celda_banco.value) 
            mes_objetivo = extraer_mes_inteligente(celda_mes.value)

            # --- PASO A: Sumar Egresos (DATA BS) ---
            suma_data_bs = 0
            for r in range(4, ws_data.max_row + 1):
                val_cta = ws_data.cell(row=r, column=13).value
                val_banco = ws_data.cell(row=r, column=10).value
                d_fecha = ws_data.cell(row=r, column=2).value
                
                d_cuenta = normalizar_texto(val_cta)
                d_banco = normalizar_texto(val_banco)
                d_monto = limpiar_venezuela(ws_data.cell(row=r, column=7).value)
                
                if d_monto != 0:
                    mes_fila = extraer_mes_inteligente(d_fecha)
                    if ("ESPECIALIZAD" in d_cuenta) and bancos_coinciden(banco_objetivo, d_banco) and mes_objetivo and mes_fila and (mes_objetivo in mes_fila or mes_fila in mes_objetivo):
                        suma_data_bs += d_monto

            # --- PASO B: Sumar Ingresos (MANEJO EXCEDENTE) desde la RAM ---
            suma_excedente = 0
            etiqueta_banco = "BP" if "PROVINCIAL" in banco_objetivo else ("BM" if "MERCANTIL" in banco_objetivo else "")
            
            if etiqueta_banco != "":
                for row_exc in excedentes_lista:
                    if len(row_exc) < 8: continue
                    
                    val_desc = row_exc[1] # Columna B (Índice 1)
                    if not val_desc: continue
                    e_desc = normalizar_texto(val_desc)
                    
                    val_mes = row_exc[2] # Columna C (Índice 2)
                    mes_exc = extraer_mes_inteligente(val_mes)
                    
                    e_desc_clean = e_desc.replace(" ", "").replace(".", "")
                    banco_match = False
                    if "PROVINCIAL" in banco_objetivo and ("BP" in e_desc_clean or "PROVINCIAL" in e_desc_clean):
                        banco_match = True
                    elif "MERCANTIL" in banco_objetivo and ("BM" in e_desc_clean or "MERCANTIL" in e_desc_clean):
                        banco_match = True

                    if ("ESPECIALIZAD" in e_desc) and banco_match and mes_objetivo and mes_exc and (mes_objetivo in mes_exc or mes_exc in mes_objetivo):
                        val_h = row_exc[7] # Columna H (Índice 7)
                        monto_h = limpiar_venezuela(val_h)
                        
                        if monto_h != 0:
                            suma_excedente += monto_h 

            # --- PASO C: Cálculo Final ---
            if suma_data_bs != 0 or suma_excedente != 0:
                resultado_final = (suma_data_bs * -1) + suma_excedente
                celda_monto.value = resultado_final
                celda_monto.number_format = '#,##0.00'
                cambios += 1
                if suma_excedente > 0: 
                    callback_log(f"   ✅ {banco_objetivo} {mes_objetivo}: {resultado_final:,.2f} (+{suma_excedente:,.2f} Exced.)")
                else: 
                    callback_log(f"   📉 {banco_objetivo} {mes_objetivo}: {resultado_final:,.2f} (Solo Data BS)")

    return cambios

# ==========================================
# ORQUESTADOR PRINCIPAL
# ==========================================
def lógica_negocio(ruta_excel, callback_log, callback_progreso):
    mensajes = []
    wb = None
    try:
        # 1. CARGA INICIAL
        callback_log("🧠 Iniciando sistemas...")
        reglas_cuentas = cargar_diccionario_cuentas() 
        reglas_areas = cargar_diccionario_areas()
        tasas_historicas = cargar_tasas_historicas(callback_log)
        
        # --- AQUÍ ESTÁ EL ARREGLO DE LA API ---
        try:
            # Conexión a la nueva API (DolarApi)
            url_api = "https://ve.dolarapi.com/v1/dolares/oficial"
            callback_log(f"🌍 Consultando Dólar Oficial en: {url_api}")
            
            resp = requests.get(url_api, timeout=10)
            data = resp.json()
            
            # Usamos 'promedio' como vimos en tu captura
            precio_dolar_hoy = float(data['promedio'])
            
            callback_log(f"💰 ¡TASA OBTENIDA!: {precio_dolar_hoy}")
        except Exception as e: 
            callback_log(f"❌ Error consultando tasa: {str(e)}")
            precio_dolar_hoy = 0
        # --------------------------------------

        callback_progreso(0.1)

        # 2. ABRIR EXCEL
        callback_log("📂 Leyendo archivo Excel...")
        wb = openpyxl.load_workbook(ruta_excel)
        callback_progreso(0.3)

        # 3. ACTUALIZAR PORTADA/HISTORICO
        if precio_dolar_hoy > 0:
            ws_portada = obtener_hoja_flexible(wb, "CUENTAS POR COBRAR")
            if ws_portada:
                ws_portada["D3"] = datetime.now().strftime("%d/%m/%Y")
                ws_portada["D4"] = precio_dolar_hoy
            ws_hist = obtener_hoja_flexible(wb, "COMPORTAMIENTO TASA")
            if ws_hist:
                fila = ws_hist.max_row + 1
                ult_fecha = ws_hist.cell(row=fila-1, column=1).value
                if ult_fecha != datetime.now().strftime("%d/%m/%Y"):
                    ws_hist.cell(row=fila, column=1, value=datetime.now().strftime("%d/%m/%Y"))
                    ws_hist.cell(row=fila, column=2, value="USD")
                    ws_hist.cell(row=fila, column=3, value=precio_dolar_hoy)
                    mensajes.append("✅ Tasa Histórica Agregada")

         # 4. CLASIFICACIÓN Y CÁLCULO USD (ESTRICTO V11)
        callback_log("🚀 Clasificando y Calculando Divisas...")
        ws_data = obtener_hoja_flexible(wb, "DATA BS")
        if ws_data:
            c_clasif, c_usd = 0, 0
            for r in range(4, ws_data.max_row + 1):
                prov = str(ws_data.cell(row=r, column=12).value).upper().strip()
                if prov:
                    if not ws_data.cell(row=r, column=13).value:
                        match = next((v for k,v in reglas_cuentas if k in prov), None)
                        if not match:
                             if "COMISION" in prov: match = "GASTOS BANCARIOS"
                             elif "IVA" in prov: match = "IMPUESTOS"
                        if match: ws_data.cell(row=r, column=13).value = match; c_clasif += 1
                    if not ws_data.cell(row=r, column=15).value:
                        match = next((v for k,v in reglas_areas if k in prov), None)
                        if match: ws_data.cell(row=r, column=15).value = match
                
                val_fecha = ws_data.cell(row=r, column=2).value
                bs = limpiar_numero(ws_data.cell(row=r, column=7).value)
                
                if val_fecha and bs != 0 and not ws_data.cell(row=r, column=8).value:
                    f_key = formatear_fecha_para_api(val_fecha)
                    tasa = buscar_tasa_inteligente(f_key, tasas_historicas)
                    if tasa and tasa > 0:
                        ws_data.cell(row=r, column=8).value = bs / tasa
                        ws_data.cell(row=r, column=8).number_format = '#,##0.00'
                        c_usd += 1

            mensajes.append(f"✅ {c_clasif} Filas Clasificadas")
            mensajes.append(f"✅ {c_usd} Conversiones a USD")
        
        callback_progreso(0.6)

       
        # 5. CONCILIACIÓN
        cambios_conc = procesar_conciliacion_compleja(wb, ruta_excel, callback_log)
        
        
        # 6. RESUMEN SEMANAL (CORREGIDO)
        #cambios_sem = procesar_resumen_semanal(wb, ruta_excel, callback_log)
        #if cambios_sem > 0: mensajes.append(f"✅ Resumen Semanal Actualizado ({cambios_sem} celdas)")
        #else: mensajes.append("ℹ️ Resumen Semanal: Sin cambios nuevos")
        
        callback_progreso(0.9)

        # 7. GUARDAR (CON ESCUDO DE SEGURIDAD)
        callback_log("💾 Guardando archivo...")
        try:
            wb.save(ruta_excel)
            # Intentamos cerrar, pero si falla el descriptor de archivo, no nos detendrá
            try:
                wb.close()
            except:
                pass 
            
            callback_progreso(1.0)
            return True, "\n".join(mensajes)

        except PermissionError:
            return False, "⚠️ CIERRA EL EXCEL. Está abierto y bloqueado."
        except Exception as e:
            # Si el error es el famoso 'Bad file descriptor', igual devolvemos True 
            # porque el archivo ya se guardó en la línea wb.save
            if "Bad file descriptor" in str(e):
                callback_progreso(1.0)
                return True, "\n".join(mensajes) + "\n⚠️ Nota: El archivo se guardó (ignora aviso de descriptor)."
            
            return False, f"❌ Error técnico: {str(e)}"

   

# ==========================================
# 🎮 INTERFAZ WEB (STREAMLIT)
# ==========================================
if __name__ == "__main__":
    st.set_page_config(page_title="Platco Bot Financiero", page_icon="💰")

    # Cargar Logo si existe
    if os.path.exists("logo.png"):
        st.image("logo.png", width=200)

    st.title("🤖 Sistema de Gestión Financiera AI")
    st.markdown("---")

    # Botón para subir archivo
    uploaded_file = st.file_uploader("📂 Sube tu archivo Excel Financiero", type=["xlsx"])

    if uploaded_file is not None:
        st.success(f"Archivo cargado: {uploaded_file.name}")
        
        if st.button("🚀 EJECUTAR AUTOMATIZACIÓN", type="primary"):
            
            # Área de logs visual
            log_container = st.empty()
            logs_historial = []

            def web_log(msg):
                logs_historial.append(f"> {msg}")
                # Muestra los últimos 10 mensajes
                log_container.text("\n".join(logs_historial[-10:]))

            bar = st.progress(0)

            try:
                web_log("--- INICIANDO PROCESO ---")
                
                # LLAMAMOS A TU LÓGICA (MODIFICADA EN EL PASO 2)
                # Nota: Pasamos 'uploaded_file' directo, no una ruta.
                archivo_resultado, texto_resultado = lógica_negocio(uploaded_file, web_log, bar.progress)
                
                web_log("--- FINALIZADO ---")
                st.success("¡Proceso Terminado!")

                # MOSTRAR LOGS COMPLETOS
                with st.expander("Ver Reporte Detallado"):
                    st.text(texto_resultado)

                # BOTÓN DE DESCARGA (VERSIÓN VPS)
                now = datetime.now().strftime("%Y%m%d_%H%M")
                with open(archivo_resultado, "rb") as f:
                    st.download_button(
                        label="📥 DESCARGAR ARCHIVO PROCESADO",
                        data=f,
                        file_name=f"Finanzas_Procesado_{now}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error(f"❌ Error Crítico: {str(e)}")

    






